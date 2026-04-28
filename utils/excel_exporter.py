from itertools import groupby

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def export_to_excel(results: list, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '키워드 분석 결과'

    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    left_mid = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BDBDBD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['블로그 주소', '오늘 방문자수', '게시글 제목', '키워드', '순위']
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 24

    rank_in_font = Font(color='1B5E20', bold=True)
    rank_out_font = Font(color='B71C1C')

    excel_row = 2
    # results는 blog_url 순서가 보장되므로 groupby로 블로그 단위 묶음
    for blog_url, blog_iter in groupby(results, key=lambda r: r['blog_url']):
        blog_rows = list(blog_iter)
        blog_start = excel_row
        visitor_text = (
            str(blog_rows[0]['visitor_count'])
            if blog_rows[0]['visitor_count'] > 0 else '-'
        )

        for i, result in enumerate(blog_rows):
            rank = result['rank']
            rank_text = f'{rank}위' if rank > 0 else '순위 밖'

            # A열: 블로그 주소 — 첫 행에만 값, 전 행에 테두리
            cell_a = ws.cell(row=excel_row, column=1)
            if i == 0:
                cell_a.value = blog_url
            cell_a.border = border

            # B열: 방문자수 — 첫 행에만 값
            cell_b = ws.cell(row=excel_row, column=2)
            if i == 0:
                cell_b.value = visitor_text
            cell_b.border = border

            # C열: 게시글 제목 — 매 행 반복
            cell_c = ws.cell(row=excel_row, column=3, value=result['post_title'])
            cell_c.border = border
            cell_c.alignment = left_mid

            # D열: 키워드
            ws.cell(row=excel_row, column=4, value=result['keyword']).border = border

            # E열: 순위
            rank_cell = ws.cell(row=excel_row, column=5, value=rank_text)
            rank_cell.alignment = center
            rank_cell.border = border
            rank_cell.font = rank_in_font if rank > 0 else rank_out_font

            excel_row += 1

        blog_end = excel_row - 1

        # 블로그 단위 셀 병합 (A, B열)
        if blog_end > blog_start:
            ws.merge_cells(f'A{blog_start}:A{blog_end}')
            ws.merge_cells(f'B{blog_start}:B{blog_end}')

        ws[f'A{blog_start}'].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        ws[f'B{blog_start}'].alignment = center

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12

    wb.save(filepath)
